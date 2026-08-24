"""
Module 01 — Literature Retrieval
v3 Implementation (Facade/Adapter pattern).

Wraps the legacy ``literature.downloader.paper_downloader.PaperDownloader``
class to satisfy the v3 ``Module01Interface`` contract.

Upstream:   none (entry point)
Downstream: 02, 03

Output files produced:
    - literature_manifest.json
    - paper_metadata.jsonl
    - download_queue.json
    - Module01_Validation_Report.md
    - module_manifest.json

v8.2.2 additions:
    - Literature registry (CSV/XLSX/JSON) with research_task_id
    - Pre-search deduplication via literature_database.json
    - Literature_Download_Report.md
    - literature_keyword_statistics.xlsx
    - Fallback query via pipeline.get_fallback()
"""

import csv
import hashlib
import importlib.util
import json
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

# ------------------------------------------------------------------ #
# Path setup: add project root for legacy ``literature`` imports
# ------------------------------------------------------------------ #
_PROJECT_ROOT = os.path.dirname(
    os.path.dirname(
        os.path.dirname(
            os.path.dirname(os.path.abspath(__file__))
        )
    )
)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# v8.2.2: Literature registry directory
_LITERATURE_DIR = Path(_PROJECT_ROOT) / "data" / "literature"
_REGISTRY_CSV = _LITERATURE_DIR / "literature_registry.csv"
_REGISTRY_XLSX = _LITERATURE_DIR / "literature_registry.xlsx"
_DATABASE_JSON = _LITERATURE_DIR / "literature_database.json"
_KEYWORD_STATS_XLSX = _LITERATURE_DIR / "literature_keyword_statistics.xlsx"
_DOWNLOAD_REPORT_MD = _LITERATURE_DIR / "Literature_Download_Report.md"

REGISTRY_FIELDS = [
    "research_task_id", "paper_id", "title", "authors", "year",
    "venue", "DOI", "arxiv_id", "keyword_source", "search_query",
    "download_source", "file_path", "hash", "status",
]

# ------------------------------------------------------------------ #
# Load the interface contract from this module's directory.
# The directory name starts with a digit (``01_…``) so it cannot be
# imported as a normal Python package; use ``importlib`` instead.
# ------------------------------------------------------------------ #
_MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
_spec = importlib.util.spec_from_file_location(
    "module_01_interface", os.path.join(_MODULE_DIR, "interface.py")
)
_interface_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_interface_mod)

Module01Interface = _interface_mod.Module01Interface
LiteratureRetrievalInput = _interface_mod.LiteratureRetrievalInput
LiteratureRetrievalOutput = _interface_mod.LiteratureRetrievalOutput

# ------------------------------------------------------------------ #
# Legacy imports
# ------------------------------------------------------------------ #
from literature.downloader.paper_downloader import PaperDownloader  # noqa: E402

from infrastructure.research_context_extractor import extract_research_context  # noqa: E402

logger = logging.getLogger(__name__)


class LiteratureRetrievalImplementation(Module01Interface):
    """Concrete v3 implementation for Module 01 — Literature Retrieval.

    This adapter wraps the legacy :class:`PaperDownloader` and exposes it
    through the v3 module lifecycle (load_config → validate_input →
    execute → validate_output → quality_assessment → write_manifest →
    write_report).
    """

    MODULE_ID = "01"
    MODULE_NAME = "Literature Retrieval"
    MODULE_VERSION = "1.0"

    # Default configuration.
    DEFAULT_CONFIG: Dict[str, Any] = {
        "max_papers": 50,
        "databases": ["arxiv", "semantic_scholar", "openreview"],
        "request_timeout": 30,
        "output_dir": "literature_output",
    }

    def __init__(self) -> None:
        self._config: Dict[str, Any] = dict(self.DEFAULT_CONFIG)
        self._downloader: Optional[PaperDownloader] = None
        self._output_dir: str = self._config["output_dir"]
        self._last_output: Optional[LiteratureRetrievalOutput] = None

    # ------------------------------------------------------------------
    # 1. load_config
    # ------------------------------------------------------------------
    def load_config(self, config: Dict[str, Any]) -> None:
        """Load and merge module-specific configuration.

        Args:
            config: Configuration dictionary. Recognised keys:
                - max_papers (int, default 50)
                - databases (List[str], default all three)
                - request_timeout (int, default 30)
                - output_dir (str, default "literature_output")
        """
        merged = dict(self.DEFAULT_CONFIG)
        merged.update(config or {})
        self._config = merged
        self._output_dir = merged.get("output_dir", "literature_output")
        self._downloader = PaperDownloader(
            request_timeout=merged.get("request_timeout", 30)
        )
        logger.info("Module 01 config loaded: %s", {k: v for k, v in merged.items() if k != "output_dir"})

    # ------------------------------------------------------------------
    # 2. validate_input
    # ------------------------------------------------------------------
    def validate_input(self, input_data: LiteratureRetrievalInput) -> bool:
        """Validate that the required ``research_task.yaml`` is present.

        Args:
            input_data: Standard module input.

        Returns:
            True if ``research_task.yaml`` exists in input_files and
            the file is readable.
        """
        if not input_data.input_files:
            logger.error("No input_files provided")
            return False
        rt_path = input_data.input_files.get("research_task.yaml")
        if not rt_path:
            logger.error("Missing required input file: research_task.yaml")
            return False
        if not os.path.exists(rt_path):
            logger.error("research_task.yaml path does not exist: %s", rt_path)
            return False
        return True

    # ------------------------------------------------------------------
    # 3. execute
    # ------------------------------------------------------------------
    def execute(self, input_data: LiteratureRetrievalInput) -> LiteratureRetrievalOutput:
        """Execute the module's core logic.

        Reads the research task, searches configured databases via the
        legacy :class:`PaperDownloader`, and writes the three core output
        files plus a manifest and validation report.

        v8.2.2: Also updates literature registry (with research_task_id),
        generates download report and keyword statistics, and queries
        fallback policy via pipeline.get_fallback().

        Args:
            input_data: Validated module input.

        Returns:
            Module output with output_files, manifest, warnings, errors.
        """
        warnings: List[str] = []
        errors: List[str] = []

        # Ensure output directory exists
        os.makedirs(self._output_dir, exist_ok=True)
        _LITERATURE_DIR.mkdir(parents=True, exist_ok=True)

        # Parse research_task.yaml
        task_config = self._parse_research_task(input_data)
        if not task_config:
            errors.append("Could not parse research_task.yaml")
            return self._build_output(
                input_data.task_id, {}, warnings, errors
            )

        # v8.2.2: Extract research_task_id
        research_task_id = task_config.get("task_id", input_data.task_id)

        # Extract research context from nested or flat fields
        ctx = extract_research_context(task_config)
        keywords = ctx["keywords"]
        max_papers = ctx["max_papers"]
        databases = ctx["databases"]
        research_question = ctx["research_question"]
        domain = ctx["domain"]
        topic = ctx["topic"]
        target = ctx["target"]
        arxiv_download_pdf = ctx["arxiv_download_pdf"]
        arxiv_prefer_latex = ctx["arxiv_prefer_latex"]

        if not keywords and research_question:
            keywords = [research_question]

        if not keywords:
            errors.append("No keywords or research_question found in research_task.yaml")
            return self._build_output(input_data.task_id, {}, warnings, errors)

        # v8.2.2: Query fallback for skill:light-literature-search
        fallback_info = self._query_skill_fallback(input_data, "light-literature-search")
        if fallback_info:
            warnings.append(fallback_info.get("message", ""))

        # v8.3: Index existing PDFs before search
        indexed_count = self._index_existing_pdfs(research_task_id)
        if indexed_count > 0:
            warnings.append(f"v8.3: Indexed {indexed_count} existing PDFs into database")

        # v8.2.2: Load existing literature database for pre-search dedup
        existing_db = self._load_literature_database()
        existing_ids: set = set()
        if existing_db:
            existing_ids = {p.get("paper_id", "") for p in existing_db.get("papers", [])}
            logger.info("Loaded %d existing papers from database for dedup", len(existing_ids))

        # v8.3: Check if we have enough papers after indexing
        total_db_papers = len(existing_db.get("papers", [])) if existing_db else 0
        if total_db_papers < 50:
            error_report = self._generate_error_report(input_data.task_id, total_db_papers)
            warnings.append(
                f"Literature count ({total_db_papers}) < 50 minimum. "
                f"Error report: {error_report}"
            )

        # Ensure downloader is initialised
        if self._downloader is None:
            self._downloader = PaperDownloader(
                request_timeout=self._config.get("request_timeout", 30)
            )

        # Search databases
        all_papers: List[Dict[str, Any]] = []
        search_queries: List[str] = []
        databases_queried: List[str] = []
        keyword_hit_map: Dict[str, int] = {}

        for keyword in keywords:
            search_queries.append(keyword)
            keyword_hit_count = 0

            if "arxiv" in databases:
                databases_queried.append("arxiv")
                try:
                    results = self._downloader.search_arxiv(keyword, max_results=max_papers)
                    self._tag_source(results, "arxiv")
                    for r in results:
                        r["_keyword_source"] = keyword
                    all_papers.extend(results)
                    keyword_hit_count += len(results)
                except Exception as exc:
                    warnings.append(f"arXiv search failed for '{keyword}': {exc}")

            if "semantic_scholar" in databases:
                if "semantic_scholar" not in databases_queried:
                    databases_queried.append("semantic_scholar")
                try:
                    results = self._downloader.search_semantic_scholar(keyword, max_results=max_papers)
                    self._tag_source(results, "semantic_scholar")
                    for r in results:
                        r["_keyword_source"] = keyword
                    all_papers.extend(results)
                    keyword_hit_count += len(results)
                except Exception as exc:
                    warnings.append(f"Semantic Scholar search failed for '{keyword}': {exc}")

            if "openreview" in databases:
                if "openreview" not in databases_queried:
                    databases_queried.append("openreview")
                try:
                    results = self._downloader.search_openreview(keyword, max_results=max_papers)
                    self._tag_source(results, "openreview")
                    for r in results:
                        r["_keyword_source"] = keyword
                    all_papers.extend(results)
                    keyword_hit_count += len(results)
                except Exception as exc:
                    warnings.append(f"OpenReview search failed for '{keyword}': {exc}")

            keyword_hit_map[keyword] = keyword_hit_count

        # Deduplicate by paper_id (keep first occurrence)
        # v8.2.2: Also check against existing database
        seen_ids: set = set()
        unique_papers: List[Dict[str, Any]] = []
        duplicates_skipped = 0
        for paper in all_papers:
            pid = paper.get("paper_id", "")
            if pid and pid not in seen_ids:
                seen_ids.add(pid)
                if pid in existing_ids:
                    paper["_status"] = "duplicate"
                    duplicates_skipped += 1
                else:
                    paper["_status"] = "pending"
                unique_papers.append(paper)
            elif pid:
                duplicates_skipped += 1

        # Truncate to max_papers
        unique_papers = unique_papers[:max_papers]

        # Build download queue (papers with a pdf_url)
        download_queue: List[Dict[str, Any]] = []
        for paper in unique_papers:
            pdf_url = paper.get("pdf_url", "")
            if pdf_url:
                download_queue.append({
                    "paper_id": paper.get("paper_id", ""),
                    "url": pdf_url,
                    "source_db": paper.get("source_db", paper.get("venue", "")),
                    "priority": 0,
                })

        # Build manifest
        manifest = {
            "total_papers": len(unique_papers),
            "total_downloadable": len(download_queue),
            "search_queries": search_queries,
            "databases_queried": databases_queried,
            "retrieval_timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "task_id": input_data.task_id,
            "research_context": {
                "domain": domain,
                "topic": topic,
                "keywords": keywords,
                "research_question": research_question,
                "target": target,
                "candidate_target": ctx["candidate_target"],
                "core_target": ctx["core_target"],
                "deep_analysis_target": ctx["deep_analysis_target"],
                "arxiv_download_pdf": arxiv_download_pdf,
                "arxiv_prefer_latex": arxiv_prefer_latex,
                "synthetic_enabled": ctx["synthetic_enabled"],
                "real_enabled": ctx["real_enabled"],
                "experiment_method": ctx["experiment_method"],
            },
        }

        # Write output files
        output_files: Dict[str, str] = {}

        # paper_metadata.jsonl
        metadata_path = os.path.join(self._output_dir, "paper_metadata.jsonl")
        self._write_jsonl(metadata_path, unique_papers)
        output_files["paper_metadata.jsonl"] = metadata_path

        # download_queue.json
        queue_path = os.path.join(self._output_dir, "download_queue.json")
        with open(queue_path, "w", encoding="utf-8") as f:
            json.dump({"queue": download_queue}, f, indent=2, ensure_ascii=False)
        output_files["download_queue.json"] = queue_path

        # literature_manifest.json
        manifest_path = os.path.join(self._output_dir, "literature_manifest.json")
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2, ensure_ascii=False)
        output_files["literature_manifest.json"] = manifest_path

        # module_manifest.json
        mod_manifest = self._generate_manifest(
            input_data.task_id, output_files, manifest, warnings, errors
        )
        mod_manifest_path = os.path.join(self._output_dir, "module_manifest.json")
        with open(mod_manifest_path, "w", encoding="utf-8") as f:
            json.dump(mod_manifest, f, indent=2, ensure_ascii=False)
        output_files["module_manifest.json"] = mod_manifest_path

        # validation report
        report = self._generate_report(input_data.task_id, output_files, manifest, warnings, errors)
        report_path = os.path.join(self._output_dir, "Module01_Validation_Report.md")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report)
        output_files["Module01_Validation_Report.md"] = report_path

        # v8.2.2: Update literature registry (CSV, XLSX, JSON) with research_task_id
        try:
            self._update_literature_registry(
                unique_papers, research_task_id, search_queries, databases_queried
            )
            output_files["literature_registry.csv"] = str(_REGISTRY_CSV)
            output_files["literature_registry.xlsx"] = str(_REGISTRY_XLSX)
            output_files["literature_database.json"] = str(_DATABASE_JSON)
        except Exception as exc:
            warnings.append(f"Registry update failed: {exc}")

        # v8.2.2: Generate Literature_Download_Report.md
        try:
            self._generate_download_report(
                input_data.task_id, research_task_id, unique_papers,
                download_queue, duplicates_skipped, keyword_hit_map,
                databases_queried, search_queries, warnings, errors,
            )
            output_files["Literature_Download_Report.md"] = str(_DOWNLOAD_REPORT_MD)
        except Exception as exc:
            warnings.append(f"Download report generation failed: {exc}")

        # v8.2.2: Generate literature_keyword_statistics.xlsx
        try:
            self._generate_keyword_statistics(keyword_hit_map, search_queries)
            output_files["literature_keyword_statistics.xlsx"] = str(_KEYWORD_STATS_XLSX)
        except Exception as exc:
            warnings.append(f"Keyword statistics generation failed: {exc}")

        # v8.3: Generate Stage_Report.md
        try:
            stage_report = self._build_stage_report(
                input_data.task_id, manifest, warnings, errors
            )
            stage_path = os.path.join(self._output_dir, "Stage_Report.md")
            with open(stage_path, "w", encoding="utf-8") as f:
                f.write(stage_report)
            output_files["Stage_Report.md"] = stage_path
        except Exception as exc:
            warnings.append(f"Stage report generation failed: {exc}")

        output = LiteratureRetrievalOutput(
            task_id=input_data.task_id,
            output_files=output_files,
            manifest=manifest,
            warnings=warnings,
            errors=errors,
        )
        self._last_output = output
        return output

    def _build_stage_report(
        self, task_id: str, manifest: Dict[str, Any],
        warnings: List[str], errors: List[str],
    ) -> str:
        """v8.3: Build Stage_Report.md for Module 01."""
        from datetime import datetime
        total = manifest.get("total_papers", 0)
        status = "完成" if not errors else "部分完成"
        lines = [
            "# Module 01 — Literature Retrieval Stage Report",
            "",
            f"- **Task ID**: {task_id}",
            f"- **时间戳**: {datetime.now().isoformat()}",
            f"- **状态**: {status}",
            "",
            "## 当前目标",
            "从多个学术数据库检索文献，构建文献数据库和索引",
            "",
            "## 输入",
            "- research_task.yaml (研究任务配置)",
            "",
            "## 输出",
            "- literature_database.json",
            "- literature_registry.csv / .xlsx",
            "- paper_metadata.jsonl",
            "- download_queue.json",
            "- Stage_Report.md",
            "",
            "## 完成状态",
            f"- 文献总数: {total}",
            f"- 数据库数量: {len(manifest.get('databases_queried', []))}",
            f"- 搜索关键词数: {len(manifest.get('search_queries', []))}",
            f"- 下载队列大小: {manifest.get('download_queue_size', 0)}",
            "",
        ]
        if warnings:
            lines.extend(["## 警告", ""])
            for w in warnings:
                lines.append(f"- {w}")
        if errors:
            lines.extend(["## 错误", ""])
            for e in errors:
                lines.append(f"- {e}")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # 4. validate_output
    # ------------------------------------------------------------------
    def validate_output(self, output: LiteratureRetrievalOutput) -> bool:
        """Validate that all required output files exist and are non-empty.

        Args:
            output: Module output to validate.

        Returns:
            True if all required files are present and valid.
        """
        required = [
            "literature_manifest.json",
            "paper_metadata.jsonl",
            "download_queue.json",
            "Module01_Validation_Report.md",
            "module_manifest.json",
        ]
        for fname in required:
            path = output.output_files.get(fname)
            if not path or not os.path.exists(path):
                logger.error("Missing or non-existent output file: %s", fname)
                return False
        return True

    # ------------------------------------------------------------------
    # 5. quality_assessment
    # ------------------------------------------------------------------
    def quality_assessment(self, output: LiteratureRetrievalOutput) -> Dict[str, Any]:
        """Assess output quality against hard requirements and soft thresholds.

        Args:
            output: Module output to assess.

        Returns:
            Dictionary with quality metrics, pass/fail status, and details.
        """
        hard_results: Dict[str, bool] = {}
        soft_results: Dict[str, Any] = {}
        issues: List[str] = []

        # Hard requirement 0: At least 1 paper metadata entry
        metadata_path = output.output_files.get("paper_metadata.jsonl", "")
        paper_count = 0
        if metadata_path and os.path.exists(metadata_path):
            with open(metadata_path, "r", encoding="utf-8") as f:
                paper_count = sum(1 for line in f if line.strip())
        hard_results["at_least_1_paper"] = paper_count >= 1
        if not hard_results["at_least_1_paper"]:
            issues.append("No paper metadata entries found")

        # Hard requirement 1: research_task.yaml has non-empty research_question
        # (already validated in execute; check manifest for search_queries)
        has_queries = bool(output.manifest.get("search_queries"))
        hard_results["has_research_question"] = has_queries
        if not has_queries:
            issues.append("No search queries derived from research_question")

        # Hard requirement 2: download_queue has at least 1 entry
        queue_path = output.output_files.get("download_queue.json", "")
        queue_count = 0
        if queue_path and os.path.exists(queue_path):
            with open(queue_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                queue_count = len(data.get("queue", []))
        hard_results["download_queue_nonempty"] = queue_count >= 1
        if not hard_results["download_queue_nonempty"]:
            issues.append("Download queue is empty")

        all_hard_pass = all(hard_results.values())

        # Soft threshold 0: Prefer >= 20 papers
        soft_results["paper_count"] = paper_count
        soft_results["paper_count_pass"] = paper_count >= 20

        # Soft threshold 1: Prefer >= 2 databases
        db_count = len(output.manifest.get("databases_queried", []))
        soft_results["database_count"] = db_count
        soft_results["database_count_pass"] = db_count >= 2

        return {
            "overall_pass": all_hard_pass,
            "hard_requirements": hard_results,
            "soft_thresholds": soft_results,
            "issues": issues,
            "paper_count": paper_count,
            "queue_count": queue_count,
        }

    # ------------------------------------------------------------------
    # 6. write_manifest
    # ------------------------------------------------------------------
    def write_manifest(self, output: LiteratureRetrievalOutput) -> Dict[str, Any]:
        """Generate the module manifest for provenance tracking.

        Args:
            output: Module output.

        Returns:
            Manifest dictionary with module metadata.
        """
        return self._generate_manifest(
            output.task_id, output.output_files, output.manifest,
            output.warnings, output.errors,
        )

    # ------------------------------------------------------------------
    # 7. write_report
    # ------------------------------------------------------------------
    def write_report(self, output: LiteratureRetrievalOutput) -> str:
        """Generate a human-readable validation report.

        Args:
            output: Module output.

        Returns:
            Markdown-formatted report string.
        """
        return self._generate_report(
            output.task_id, output.output_files, output.manifest,
            output.warnings, output.errors,
        )

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _tag_source(papers: List[Dict[str, Any]], source: str) -> None:
        """Tag each paper dict with a ``source_db`` field."""
        for p in papers:
            if "source_db" not in p:
                p["source_db"] = source

    def _parse_research_task(self, input_data: LiteratureRetrievalInput) -> Dict[str, Any]:
        """Parse research_task.yaml from the input files.

        Tries YAML first, falls back to JSON.
        """
        rt_path = input_data.input_files.get("research_task.yaml", "")
        if not rt_path or not os.path.exists(rt_path):
            # Try config dict from input_data
            if input_data.config:
                return input_data.config
            return {}

        try:
            import yaml
            with open(rt_path, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
            return data if isinstance(data, dict) else {}
        except ImportError:
            pass
        except Exception as exc:
            logger.warning("YAML parse failed: %s; trying JSON", exc)

        try:
            with open(rt_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    @staticmethod
    def _write_jsonl(path: str, records: List[Dict[str, Any]]) -> None:
        """Write a list of dicts as JSONL."""
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            for record in records:
                f.write(json.dumps(record, ensure_ascii=False) + "\n")

    def _build_output(
        self,
        task_id: str,
        output_files: Dict[str, str],
        manifest: Dict[str, Any],
        warnings: List[str],
        errors: List[str],
    ) -> LiteratureRetrievalOutput:
        """Construct a LiteratureRetrievalOutput dataclass."""
        return LiteratureRetrievalOutput(
            task_id=task_id,
            output_files=output_files,
            manifest=manifest,
            warnings=warnings,
            errors=errors,
        )

    def _generate_manifest(
        self,
        task_id: str,
        output_files: Dict[str, str],
        manifest: Dict[str, Any],
        warnings: List[str],
        errors: List[str],
    ) -> Dict[str, Any]:
        """Build the module manifest dictionary."""
        return {
            "module_id": self.MODULE_ID,
            "module_name": self.MODULE_NAME,
            "module_version": self.MODULE_VERSION,
            "task_id": task_id,
            "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "input_files": {},
            "output_files": output_files,
            "manifest_data": manifest,
            "warnings": warnings,
            "errors": errors,
            "status": "COMPLETED" if not errors else "FAILED",
        }

    def _generate_report(
        self,
        task_id: str,
        output_files: Dict[str, str],
        manifest: Dict[str, Any],
        warnings: List[str],
        errors: List[str],
    ) -> str:
        """Build a Markdown validation report."""
        lines = [
            f"# Module 01 — Literature Retrieval Validation Report",
            "",
            f"**Task ID:** {task_id}",
            f"**Timestamp:** {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}",
            f"**Status:** {'COMPLETED' if not errors else 'FAILED'}",
            "",
            "## Summary",
            "",
            f"- Total papers found: {manifest.get('total_papers', 0)}",
            f"- Downloadable papers: {manifest.get('total_downloadable', 0)}",
            f"- Databases queried: {', '.join(manifest.get('databases_queried', []))}",
            f"- Search queries: {len(manifest.get('search_queries', []))}",
            "",
            "## Output Files",
            "",
        ]
        for fname, fpath in output_files.items():
            exists = "YES" if os.path.exists(fpath) else "NO"
            lines.append(f"- `{fname}` — exists: {exists}")
        if warnings:
            lines.append("")
            lines.append("## Warnings")
            lines.append("")
            for w in warnings:
                lines.append(f"- {w}")
        if errors:
            lines.append("")
            lines.append("## Errors")
            lines.append("")
            for e in errors:
                lines.append(f"- {e}")
        return "\n".join(lines) + "\n"

    # ------------------------------------------------------------------
    # v8.3: Index existing PDFs from data/literature/pdf/
    # ------------------------------------------------------------------

    def _index_existing_pdfs(self, research_task_id: str) -> int:
        """v8.3: Scan existing PDFs and index them into literature_database.json.

        This method scans data/literature/pdf/ for PDF files that were
        downloaded but not yet indexed, and adds them to the database
        with proper metadata.

        Returns:
            Number of newly indexed papers.
        """
        pdf_dir = _LITERATURE_DIR / "pdf"
        if not pdf_dir.exists():
            return 0

        # Load existing database
        db = self._load_literature_database()
        db_papers = db.get("papers", [])
        db_ids = {p.get("paper_id", "") for p in db_papers}

        # Load existing registry
        existing_entries: List[Dict[str, Any]] = []
        if _REGISTRY_CSV.exists():
            try:
                with open(_REGISTRY_CSV, "r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    existing_entries = list(reader)
            except Exception:
                pass
        existing_reg_ids = {e.get("paper_id", "") for e in existing_entries}

        new_count = 0
        for pdf_file in sorted(pdf_dir.glob("*.pdf")):
            filename = pdf_file.stem  # e.g., "2401.00001"

            # Extract arXiv ID from filename
            arxiv_id = filename
            paper_id = f"arxiv_{arxiv_id}"

            if paper_id in db_ids:
                continue

            # Try to extract metadata from filename
            year = ""
            try:
                if len(arxiv_id) >= 4:
                    yy = int(arxiv_id[:2])
                    year = 2000 + yy if yy < 50 else 1900 + yy
            except (ValueError, IndexError):
                pass

            # Compute file hash
            file_hash = ""
            try:
                with open(pdf_file, "rb") as f:
                    file_hash = hashlib.md5(f.read()).hexdigest()
            except Exception:
                pass

            paper_entry = {
                "research_task_id": research_task_id,
                "paper_id": paper_id,
                "title": "",  # Will be filled by Module 02/03
                "authors": "",
                "year": year,
                "venue": "arXiv",
                "doi": "",
                "arxiv_id": arxiv_id,
                "source_db": "arxiv",
                "pdf_url": f"https://arxiv.org/abs/{arxiv_id}",
                "pdf_path": str(pdf_file),
                "latex_path": "",
                "markdown_path": "",
                "figures_path": "",
                "status": "downloaded",
                "file_hash": file_hash,
                "added_timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            }
            db_papers.append(paper_entry)
            db_ids.add(paper_id)
            new_count += 1

            # Also add to registry
            if paper_id not in existing_reg_ids:
                reg_entry = {
                    "research_task_id": research_task_id,
                    "paper_id": paper_id,
                    "title": "",
                    "authors": "",
                    "year": year,
                    "venue": "arXiv",
                    "DOI": "",
                    "arxiv_id": arxiv_id,
                    "keyword_source": "pdf_scan",
                    "search_query": "",
                    "download_source": "arxiv",
                    "file_path": str(pdf_file),
                    "hash": file_hash,
                    "status": "downloaded",
                }
                existing_entries.append(reg_entry)
                existing_reg_ids.add(paper_id)

        # Write updated database
        db["papers"] = db_papers
        db["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        db["total_papers"] = len(db_papers)
        with open(_DATABASE_JSON, "w", encoding="utf-8") as f:
            json.dump(db, f, indent=2, ensure_ascii=False)

        # Write updated registry CSV
        with open(_REGISTRY_CSV, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=REGISTRY_FIELDS)
            writer.writeheader()
            for entry in existing_entries:
                writer.writerow({k: entry.get(k, "") for k in REGISTRY_FIELDS})

        # Write XLSX
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Literature Registry"
            ws.append(REGISTRY_FIELDS)
            for entry in existing_entries:
                ws.append([entry.get(k, "") for k in REGISTRY_FIELDS])
            wb.save(str(_REGISTRY_XLSX))
        except ImportError:
            logger.warning("openpyxl not available, skipping XLSX registry update")

        logger.info("v8.3: Indexed %d new PDFs from existing files", new_count)
        return new_count

    def _generate_error_report(self, task_id: str, paper_count: int, min_required: int = 50) -> str:
        """v8.3: Generate error report when literature is insufficient."""
        report_path = str(_LITERATURE_DIR / "Literature_Insufficient_Error.md")
        lines = [
            "# Literature Insufficiency Error Report",
            "",
            f"**Task ID:** {task_id}",
            f"**Timestamp:** {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}",
            "",
            "## 问题描述",
            "",
            f"当前文献数量: {paper_count}",
            f"最低要求: {min_required}",
            "",
            "## 影响",
            "",
            "文献数量不足，无法进入Literature Intelligence阶段。",
            "需要补充更多论文到 data/literature/pdf/ 目录。",
            "",
            "## 补充方式",
            "",
            "1. 手动下载相关论文PDF到 data/literature/pdf/",
            "2. 重新运行 Module 01 进行索引",
            "3. 或通过 arXiv API 自动检索下载",
            "",
        ]
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        return report_path

    # ------------------------------------------------------------------
    # v8.2.2: Literature Registry and Fallback Methods
    # ------------------------------------------------------------------

    def _query_skill_fallback(
        self, input_data: LiteratureRetrievalInput, skill_name: str
    ) -> Optional[Dict[str, Any]]:
        """Query fallback policy via pipeline.get_fallback().

        Modules MUST NOT decide fallback on their own — they query
        the pipeline which reads from dependency_policy.yaml.

        Returns None if pipeline reference is not available (e.g., standalone run).
        """
        pipeline = input_data.context.get("pipeline") if input_data.context else None
        if pipeline is None:
            return None

        dependency_type = f"skill:{skill_name}"
        try:
            fallback = pipeline.get_fallback("01", dependency_type)
            if fallback.get("action") == "block":
                logger.warning("Fallback blocked in %s mode: %s",
                               getattr(pipeline, "run_mode", "unknown"),
                               fallback.get("reason", ""))
            elif fallback.get("action") != "none":
                logger.info("Fallback policy for %s: action=%s, message=%s",
                            dependency_type, fallback.get("action"), fallback.get("message", ""))
            return fallback
        except Exception as exc:
            logger.warning("Fallback query failed for %s: %s", dependency_type, exc)
            return None

    def _load_literature_database(self) -> Dict[str, Any]:
        """Load existing literature_database.json for deduplication."""
        if not _DATABASE_JSON.exists():
            return {}
        try:
            with open(_DATABASE_JSON, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as exc:
            logger.warning("Failed to load literature database: %s", exc)
            return {}

    def _update_literature_registry(
        self,
        papers: List[Dict[str, Any]],
        research_task_id: str,
        search_queries: List[str],
        databases_queried: List[str],
    ) -> None:
        """Update literature registry (CSV, XLSX, JSON) with research_task_id.

        Appends new papers to existing registry; marks duplicates.
        """
        # Load existing registry entries
        existing_entries: List[Dict[str, Any]] = []
        if _REGISTRY_CSV.exists():
            try:
                with open(_REGISTRY_CSV, "r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    existing_entries = list(reader)
            except Exception:
                pass

        existing_ids = {e.get("paper_id", "") for e in existing_entries}

        # Build new entries
        new_entries: List[Dict[str, Any]] = []
        for paper in papers:
            pid = paper.get("paper_id", "")
            if pid in existing_ids:
                continue
            entry = {
                "research_task_id": research_task_id,
                "paper_id": pid,
                "title": paper.get("title", ""),
                "authors": paper.get("authors", "") if isinstance(paper.get("authors", ""), str)
                          else ", ".join(paper.get("authors", [])),
                "year": paper.get("year", ""),
                "venue": paper.get("venue", ""),
                "DOI": paper.get("doi", paper.get("DOI", "")),
                "arxiv_id": paper.get("arxiv_id", ""),
                "keyword_source": paper.get("_keyword_source", ""),
                "search_query": paper.get("_keyword_source", ""),
                "download_source": paper.get("source_db", ""),
                "file_path": "",
                "hash": "",
                "status": paper.get("_status", "pending"),
            }
            new_entries.append(entry)
            existing_ids.add(pid)

        all_entries = existing_entries + new_entries

        # Write CSV
        with open(_REGISTRY_CSV, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=REGISTRY_FIELDS)
            writer.writeheader()
            for entry in all_entries:
                writer.writerow({k: entry.get(k, "") for k in REGISTRY_FIELDS})

        # Write XLSX
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Literature Registry"
            ws.append(REGISTRY_FIELDS)
            for entry in all_entries:
                ws.append([entry.get(k, "") for k in REGISTRY_FIELDS])
            wb.save(str(_REGISTRY_XLSX))
        except ImportError:
            logger.warning("openpyxl not available, skipping XLSX registry")

        # Update JSON database
        db = self._load_literature_database()
        db_papers = db.get("papers", [])
        db_ids = {p.get("paper_id", "") for p in db_papers}
        for paper in papers:
            pid = paper.get("paper_id", "")
            if pid and pid not in db_ids:
                db_papers.append({
                    "research_task_id": research_task_id,
                    "paper_id": pid,
                    "title": paper.get("title", ""),
                    "authors": paper.get("authors", ""),
                    "year": paper.get("year", ""),
                    "venue": paper.get("venue", ""),
                    "doi": paper.get("doi", ""),
                    "arxiv_id": paper.get("arxiv_id", ""),
                    "source_db": paper.get("source_db", ""),
                    "pdf_url": paper.get("pdf_url", ""),
                    "status": paper.get("_status", "pending"),
                    "added_timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                })
                db_ids.add(pid)

        db["papers"] = db_papers
        db["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        db["total_papers"] = len(db_papers)
        with open(_DATABASE_JSON, "w", encoding="utf-8") as f:
            json.dump(db, f, indent=2, ensure_ascii=False)

        logger.info("Registry updated: %d new entries, %d total",
                     len(new_entries), len(all_entries))

    def _generate_download_report(
        self,
        task_id: str,
        research_task_id: str,
        papers: List[Dict[str, Any]],
        download_queue: List[Dict[str, Any]],
        duplicates_skipped: int,
        keyword_hit_map: Dict[str, int],
        databases_queried: List[str],
        search_queries: List[str],
        warnings: List[str],
        errors: List[str],
    ) -> None:
        """Generate Literature_Download_Report.md."""
        new_papers = [p for p in papers if p.get("_status") != "duplicate"]
        dup_papers = [p for p in papers if p.get("_status") == "duplicate"]

        lines = [
            "# Literature Download Report",
            "",
            f"**Task ID:** {task_id}",
            f"**Research Task ID:** {research_task_id}",
            f"**Timestamp:** {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}",
            "",
            "## Summary",
            "",
            f"- Total papers found: {len(papers)}",
            f"- New papers (pending download): {len(new_papers)}",
            f"- Duplicates (already in registry): {len(dup_papers)}",
            f"- Duplicates skipped during search: {duplicates_skipped}",
            f"- Downloadable papers (with PDF URL): {len(download_queue)}",
            f"- Databases queried: {', '.join(databases_queried)}",
            f"- Search queries: {len(search_queries)}",
            "",
            "## Keyword Hit Statistics",
            "",
            "| Keyword | Papers Found |",
            "|---------|-------------|",
        ]
        for kw, count in keyword_hit_map.items():
            lines.append(f"| {kw} | {count} |")

        lines.extend(["", "## New Papers (Pending Download)", ""])
        lines.append("| # | Paper ID | Title | Source | Year |")
        lines.append("|---|----------|-------|--------|------|")
        for i, p in enumerate(new_papers[:50], 1):
            title = (p.get("title", "")[:60] + "...") if len(p.get("title", "")) > 60 else p.get("title", "")
            lines.append(f"| {i} | {p.get('paper_id', '')} | {title} | {p.get('source_db', '')} | {p.get('year', '')} |")
        if len(new_papers) > 50:
            lines.append(f"| ... | ({len(new_papers) - 50} more) | | | |")

        if dup_papers:
            lines.extend(["", "## Duplicate Papers (Already in Registry)", ""])
            lines.append("| # | Paper ID | Title |")
            lines.append("|---|----------|-------|")
            for i, p in enumerate(dup_papers[:20], 1):
                title = (p.get("title", "")[:60] + "...") if len(p.get("title", "")) > 60 else p.get("title", "")
                lines.append(f"| {i} | {p.get('paper_id', '')} | {title} |")

        if warnings:
            lines.extend(["", "## Warnings", ""])
            for w in warnings:
                lines.append(f"- {w}")
        if errors:
            lines.extend(["", "## Errors", ""])
            for e in errors:
                lines.append(f"- {e}")

        with open(_DOWNLOAD_REPORT_MD, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")

    def _generate_keyword_statistics(
        self,
        keyword_hit_map: Dict[str, int],
        search_queries: List[str],
    ) -> None:
        """Generate literature_keyword_statistics.xlsx."""
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Keyword Statistics"
            ws.append(["Keyword", "Papers Found", "Percentage"])
            total = sum(keyword_hit_map.values()) or 1
            for kw, count in keyword_hit_map.items():
                pct = f"{count / total * 100:.1f}%"
                ws.append([kw, count, pct])
            ws.append(["TOTAL", total, "100.0%"])

            # Bold the total row
            from openpyxl.styles import Font
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

            wb.save(str(_KEYWORD_STATS_XLSX))
        except ImportError:
            logger.warning("openpyxl not available, skipping keyword statistics XLSX")



